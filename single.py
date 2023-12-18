from openpyxl import load_workbook
from ortools.linear_solver import pywraplp
from itertools import combinations
import csv

def load_sheet(sheet, data_dict, NoneValue= None):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) for cell in row])
    
    row_headers = [row[0] for row in data[1:] if row[0] != 'None']
    col_headers = [x for x in data[0][1:] if x != 'None']

    for i in range(len(row_headers)):
        for j in range(len(col_headers)):
            if data[i + 1][j + 1] != 'None':
                data_dict[row_headers[i], col_headers[j]] = int(data[i + 1][j + 1])
            else:
                data_dict[row_headers[i], col_headers[j]] = NoneValue

def load_sheet_3(sheet, data_dict, NoneValue= None):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) for cell in row])
    
    row_headers = [row[0] for row in data[1:] if row[0] != 'None']
    col_headers = [x for x in data[0][1:] if x != 'None']

    for i in range(len(row_headers)):
        for j in range(len(col_headers)):
            if data[i + 1][j + 1] != 'None':
                data_dict[row_headers[i], col_headers[j]] = float(data[i + 1][j + 1])
            else:
                data_dict[row_headers[i], col_headers[j]] = NoneValue

def task1():
    solver = pywraplp.Solver.CreateSolver("GLOP_LINEAR_PROGRAMMING")

    # 1. Load the xlsx file
    xlsx_file = load_workbook('Assignment_DA_2_Task_1_data.xlsx', read_only= True)

    # define the sheets
    supplier_stock_sheet = xlsx_file['Supplier stock']
    raw_mat_costs_sheet = xlsx_file['Raw material costs']
    raw_mat_ship_sheet = xlsx_file['Raw material shipping']
    prod_req_sheet = xlsx_file['Product requirements']
    prod_cap_sheet = xlsx_file['Production capacity']
    prod_cost_sheet = xlsx_file['Production cost']
    customer_demand_sheet = xlsx_file['Customer demand']
    ship_cost_sheet = xlsx_file['Shipping costs']

    # load sheets
    supplier_stock = {}
    load_sheet(supplier_stock_sheet, supplier_stock, NoneValue= 0)
    raw_mat_costs = {}
    load_sheet(raw_mat_costs_sheet, raw_mat_costs, NoneValue= 0)
    raw_mat_ship = {}
    load_sheet(raw_mat_ship_sheet, raw_mat_ship, NoneValue= 0)
    prod_req = {}
    load_sheet(prod_req_sheet, prod_req, NoneValue= 0)
    prod_cap = {}
    load_sheet(prod_cap_sheet, prod_cap, NoneValue= 0)
    prod_cost = {}
    load_sheet(prod_cost_sheet, prod_cost, NoneValue= 0)
    customer_demand = {}
    load_sheet(customer_demand_sheet, customer_demand, NoneValue= 0)
    ship_cost = {}
    load_sheet(ship_cost_sheet, ship_cost, NoneValue= 0)

    suppliers = sorted(list(set([key[0] for key in supplier_stock.keys()])))
    materials = sorted(list(set([key[1] for key in supplier_stock.keys()])))
    factories = sorted(list(set([key[1] for key in raw_mat_ship.keys()])))
    products = sorted(list(set([key[0] for key in prod_req.keys()])))
    customers = sorted(list(set([key[1] for key in customer_demand.keys()])))

    # 2. Define the variables
    # supplier order variables
    supplier_order_vars = {}
    for supplier in suppliers:
        for material in materials:
            for factory in factories:
                supplier_order_vars[supplier, material, factory] = solver.IntVar(0, solver.infinity(), f'supplier_order[{supplier}, {material}, {factory}]')
    
    # production volume variables
    production_vol_vars = {}
    for product in products:
        for factory in factories:
            production_vol_vars[product, factory] = solver.IntVar(0, solver.infinity(), f'production_volume[{product}, {factory}]')
    
    # customer delivery variables
    customer_delivery_vars = {}
    for customer in customers:
        for product in products:
            for factory in factories:
                customer_delivery_vars[customer, product, factory] =\
                    solver.IntVar(0, solver.infinity(), f'customer_delivery[{customer}, {product}, {factory}]')

    # 3. Define and implement the constraints that ensure factories produce more than they ship to the customers
    for product in products:
        for factory in factories:
            solver.Add(production_vol_vars[product, factory] - sum([customer_delivery_vars[customer, product, factory] for customer in customers]) >= 0)
    
    # 4. Define and implement the constraints that ensure that customer demand is met
    for customer in customers:
        for product in products:
            solver.Add(sum([customer_delivery_vars[customer, product, factory] for factory in factories]) >= customer_demand[product, customer])

    # 5. Define and implement the constraints that ensure that suppliers have all ordered items in stock
    for supplier in suppliers:
        for material in materials:
            solver.Add(sum([supplier_order_vars[supplier, material, factory] for factory in factories]) <= supplier_stock[supplier, material])

    for material in materials:
        for factory in factories:
            solver.Add(sum([supplier_order_vars[supplier, material, factory] for supplier in suppliers]) -\
                       sum([production_vol_vars[product, factory] * prod_req[product, material] for product in products]) >= 0)
    
    # 6. Define and implement the constraints that ensure that the manufacturing capacities are not exceeded
    for product in products:
        for factory in factories:
            solver.Add(production_vol_vars[product, factory] <= prod_cap[product, factory])
    
    # 7. Define and implement the objective function.
    supplier_cost = sum([supplier_order_vars[supplier, material, factory] * raw_mat_costs[supplier, material]\
                         for supplier in suppliers for material in materials for factory in factories])
    
    supplier_ship_cost = sum([supplier_order_vars[supplier, material, factory] * raw_mat_ship[supplier, factory]\
                         for supplier in suppliers for material in materials for factory in factories])

    production_cost = sum([production_vol_vars[product, factory] * prod_cost[product, factory]\
                           for product in products for factory in factories])
    
    customer_ship_cost = sum([customer_delivery_vars[customer, product, factory] * ship_cost[factory, customer]\
                              for customer in customers for product in products for factory in factories])
    
    overall_cost = supplier_cost + supplier_ship_cost + production_cost + customer_ship_cost

    solver.Minimize(overall_cost)

    # 8. Solve the linear program and determine the optimal overall cost 
    status = solver.Solve()

    if status == pywraplp.Solver.OPTIMAL:
        print(f"Overall Cost: {solver.Objective().Value()}")

        print("\n")
    # 9. Determine for each factory how much material has to be ordered from each individual supplier
        for factory in factories:
            for material in materials:
                for supplier in suppliers:
                    if int(round(supplier_order_vars[supplier, material, factory].solution_value())) != 0:
                        print(f"{factory} orders {material} from {supplier} for {int(round(supplier_order_vars[supplier, material, factory].solution_value()))}")
        
        print("\n")
    # 10. Determine for each factory what the supplier bill comprising material cost and delivery will be for each supplier
        for factory in factories:
            for supplier in suppliers:
                billing = int(round(sum([supplier_order_vars[supplier, material, factory].solution_value() * raw_mat_costs[supplier, material] for material in materials]) +\
                        sum([supplier_order_vars[supplier, material, factory].solution_value() *  raw_mat_ship[supplier, factory] for material in materials])))
                if billing != 0:
                    print(f"For {factory}, {supplier} bills {billing}")

    # 11. Determine for each factory how many units of each product are being manufactured. 
    # Also determine the total manufacturing cost for each individual factory.
        print("\n")
        for factory in factories:
            for product in products:
                if int(round(production_vol_vars[product, factory].solution_value())) != 0:
                    print(f"{factory} manufactured {product} for {int(round(production_vol_vars[product, factory].solution_value()))}")
            
            manu_cost = int(round(sum([production_vol_vars[product, factory].solution_value() * prod_cost[product, factory]\
                                        for product in products])))
            print(f"Overall manufacturing cost of {factory} is {manu_cost}")
    
    # 12. Determine for each customer how many units of each product are being shipped from eachfactory
    # Also determine the total shipping cost per customer.
        print("\n")
        for customer in customers:
            for product in products:
                for factory in factories:
                    if int(round(customer_delivery_vars[customer, product, factory].solution_value())) != 0:
                        print(f"To {customer}, {int(round(customer_delivery_vars[customer, product, factory].solution_value()))} of {product} are shipped from {factory}")
            
            total_ship_cost = int(round(sum(
                [customer_delivery_vars[customer, product, factory].solution_value() * ship_cost[factory, customer]\
                 for product in products for factory in factories]
            )))
            print(f'Total Shipping Cost for {customer} is {total_ship_cost}')
    
    # 13. Determine for each customer the fraction of each material each factory has to order for manufacturing products delivered to that particular customer. 
    # Based on this calculate the overall unit cost of each product per customer including the raw materials used for the manufacturing of the customer’s specific product, 
    # the cost of manufacturing for the specific customer and all relevant shipping costs.
        print("\n")
        unit_mat_cost = {}
        for factory in factories:
            for material in materials:
                total_mat_cost = int(round(sum([supplier_order_vars[supplier, material, factory].solution_value() * raw_mat_costs[supplier, material]\
                                                          for supplier in suppliers]) +\
                                                    sum([supplier_order_vars[supplier, material, factory].solution_value() * raw_mat_ship[supplier, factory]\
                                                          for supplier in suppliers])))
                total_mat_amount = int(round(sum([supplier_order_vars[supplier, material, factory].solution_value()\
                                                                            for supplier in suppliers])))
                unit_mat_cost[material, factory] = float(total_mat_cost) / float(total_mat_amount) if total_mat_amount != 0 else 0
                
        for customer in customers:
            for product in products:
                for factory in factories:
                    if int(round(customer_delivery_vars[customer, product, factory].solution_value())) != 0:
                        prod_amount = int(round(customer_delivery_vars[customer, product, factory].solution_value()))
                        for material in materials:
                            if prod_req[product, material] != 0:
                                mat_amount = prod_amount * prod_req[product, material]
                                print(f"To {customer}, to deliver {prod_amount} {product}, {factory} orders {material} for {mat_amount}")
        
        print("\n")
        for customer in customers:
            for product in products:
                product_count = int(round(sum(customer_delivery_vars[customer, product, factory].solution_value() for factory in factories)))
                if product_count != 0:
                    total_mat_cost = sum(customer_delivery_vars[customer, product, factory].solution_value() * prod_req[product, material] * unit_mat_cost[material, factory]\
                                         for factory in factories for material in materials)
                    total_prod_cost = sum(customer_delivery_vars[customer, product, factory].solution_value() * prod_cost[product, factory] \
                                          for factory in factories) +\
                                        sum(customer_delivery_vars[customer, product, factory].solution_value() * ship_cost[factory, customer] \
                                            for factory in factories)
                    total_cost = total_mat_cost + total_prod_cost
                    print(f"{customer}, {product}: {total_cost/product_count:.2f}")

    else:
        print("The problem does not have an optimal solution.")

def task2():

    solver = pywraplp.Solver.CreateSolver("CBC_MIXED_INTEGER_PROGRAMMING")
    
    # 1. Load the xlsx file
    xlsx_file = load_workbook('Assignment_DA_2_Task_2_data.xlsx', read_only= True)
    distance_sheet = xlsx_file['Distances']
    distance = {}
    load_sheet(distance_sheet, distance)
    
    towns_to_visit = ['Cork', 'Dublin', 'Limerick', 'Waterford', 'Galway', 'Wexford', 'Belfast', 'Athlone', 'Rosslare', 'Wicklow']

    # 2. For each pair of towns that need to be visited create a decision variable to decide if this leg should be included into the route
    legs = {}

    for town1 in towns_to_visit:
        for town2 in towns_to_visit:
            if town1 != town2:
                legs[town1, town2] = solver.IntVar(0, 1, "")
    
    for town in towns_to_visit:
    # 3. Define and implement the constraints that ensure that the delivery driver arrives in each of the towns that need to be visited
        solver.Add(sum(legs[town, town2] for town2 in towns_to_visit if town2 != town) == 1)
    
    # 4. Define and implement the constraints that ensure that the driver departs each of the towns that need to be visited
        solver.Add(sum(legs[town1, town] for town1 in towns_to_visit if town != town1) == 1)
    
    # 5. Define and implement the constraints that ensure that there are no disconnected selfcontained circles in the route
    subtowns = [subtown for i in range(2, len(towns_to_visit)) for subtown in combinations(towns_to_visit, i)]
        
    for subtown in subtowns:
        solver.Add(sum(legs[town1, town2] for town1 in subtown for town2 in subtown if town1 != town2) <= len(subtown) - 1)
    
    # 6. Define and implement the objective function to minimise the overall distance travelled.
    overall_distance = sum(legs[town1, town2] * distance[town1, town2] for town1 in towns_to_visit for town2 in towns_to_visit if town1 != town2)
    solver.Minimize(overall_distance)

    status = solver.Solve()

    if status == pywraplp.Solver.OPTIMAL:
        print(f"Overall Distance: {solver.Objective().Value()}")
        print("\n")

        current_town = 'Cork'
        while True:
            next_town = next(town2 for town2 in towns_to_visit if town2 != current_town and legs[current_town, town2].solution_value() > 0)
            print(current_town, '->', next_town, ':', distance[current_town, next_town])
            current_town = next_town

            if current_town == 'Cork':
                break
    else:
        print('The problem does not have an optimal solution')

def task3(currency= 'USD'):

    # 1. Load the xlsx file
    xlsx_file = load_workbook('Assignment_DA_2_Task_3_data.xlsx', read_only= True)

    usd_sheet = xlsx_file['USD']
    eur_sheet = xlsx_file['EUR']
    currency_sheet = xlsx_file['Currency']

    usd_data = {}
    load_sheet_3(usd_sheet, usd_data)

    eur_data = {}
    load_sheet_3(eur_sheet, eur_data)

    currency_data = {}
    load_sheet_3(currency_sheet, currency_data)

    timestamps = sorted(list(set([key[0] for key in usd_data.keys()])))
    stocks = list(set([key[1] for key in usd_data.keys()])) + list(set([key[1] for key in eur_data.keys()]))

    # Create a parameter to decide which currency to use
    if currency == 'USD':
        for key, value in eur_data.items():
            eur_data[key] = currency_data[key[0], 'EURUSD'] * value
    else:
        for key, value in usd_data.items():
            usd_data[key] = value / currency_data[key[0], 'EURUSD']

    stocks_data = {}
    for key, value in eur_data.items():
        stocks_data[key] = value
    for key, value in usd_data.items():
        stocks_data[key] = value
    
    # calculate the monthly return
    return_data = {}
    for i in range(1, len(timestamps)):
        for stock in stocks:
            return_data[timestamps[i], stock] = \
                stocks_data[timestamps[i], stock] / stocks_data[timestamps[i - 1], stock]

    print(f"task3_A_{currency}")

    # Determine and output the overall average monthly reward for each investment position
    average_reward_data = {}
    for stock in stocks:
        average_reward_data[stock] = sum([return_data[timestamp, stock] for timestamp in timestamps[1:]]) / (len(timestamps) - 1)
        print(f"The overall average monthly reward of {stock} is {average_reward_data[stock]}")

    # 2. Create a Linear Program to determine the reward that optimal timing the market could have
    #achieved over the past five years using the OR Tools wrapper of the GLOP_LINEAR_PROGRAMMING solver

    solver1 = pywraplp.Solver.CreateSolver("GLOP_LINEAR_PROGRAMMING")

    # For each month create decision variables that indicate the percentage of each position held as well as 
    # the percentage of cash not invested during this month 
    percent_var = {}
    positions = stocks + ['Cash']
    for timestamp in timestamps:
        for position in positions:
            percent_var[timestamp, position] = solver1.NumVar(0, 1, "")
    
    # Identify and create the implicit constraints to ensure that the investment portfolio always adds up to 100%
    for timestamp in timestamps:
        solver1.Add(sum(percent_var[timestamp, position] for position in positions) == 1.0)

    # Investing everything into one single position is not good practice. 
    # Therefore, identify and create constraints that ensure that no single investment position is ever more than 30% of the overall portfolio
    for timestamp in timestamps:
        for position in positions:
            solver1.Add(percent_var[timestamp, position] <= 0.3)
    
    # Identify and implement an objective function that maximises the overall reward of the portfolio by summing all respective monthly returns
    return_vars = {}
    for timestamp in timestamps[1:]:
        return_vars[timestamp] = sum(percent_var[timestamp, stock] * return_data[timestamp, stock] for stock in stocks)

    solver1.Maximize(sum(return_vars[timestamp] for timestamp in timestamps[1:]))

    status = solver1.Solve()

    if status == pywraplp.Solver.OPTIMAL:
        print("Task3_B", currency)
        with open(f"task3_B_{currency}.csv", "a", newline= "") as f:
            csvwriter = csv.writer(f)
            csvwriter.writerow([None] + positions)
            for timestamp in timestamps:
                csvwriter.writerow([timestamp[:7]] + [round(percent_var[timestamp, position].solution_value() * 100.0, 2) for position in positions])
            
        print("Overall Average Monthly Reward: ", solver1.Objective().Value() / (len(timestamps) - 1))



    # 3. Create another Linear Program to determine such an optimal portfolio that minimises the investment risk

    solver2 = pywraplp.Solver.CreateSolver("GLOP_LINEAR_PROGRAMMING")

    # Create decision variables that indicate the percentage of each position held in the portfolio during the entire investment period
    portfolio_vars = {}
    for timestamp in timestamps:
        for stock in stocks:
            portfolio_vars[timestamp, stock] = solver2.NumVar(0, 1, "")
    
    # Create the implicit constraint that the investment portfolio always adds up to 100%
    for timestamp in timestamps:
        solver2.Add(sum(portfolio_vars[timestamp, stock] for stock in stocks) == 1.0)
    
    # Identify and create constraints to ensure that no single investment position is ever more than 30% of the overall portfolio
    for timestamp in timestamps:
        for stock in stocks:
            solver2.Add(portfolio_vars[timestamp, stock] <= 0.3)
    
    # Create a constraint to ensure that the overall average monthly reward of the portfolio is 
    # at least 0.5% over the five-year investment period
    solver2.Add(sum(sum(portfolio_vars[timestamp, stock] * return_data[timestamp, stock] for stock in stocks) for timestamp in timestamps[1:])\
                    / (len(timestamps) - 1) >= 1.005)

    # Create these additional variables
    bounce_vars = {}
    for timestamp in timestamps[1:]:
        bounce_vars[timestamp] = solver2.NumVar(0, solver2.infinity(), "")
    
    # implement the necessary constraints for bounding the deviation
    for timestamp in timestamps[1:]:
        solver2.Add(sum(portfolio_vars[timestamp, stock] * (return_data[timestamp, stock] - average_reward_data[stock]) for stock in stocks) \
                    >= -bounce_vars[timestamp])
        solver2.Add(sum(portfolio_vars[timestamp, stock] * (return_data[timestamp, stock] - average_reward_data[stock]) for stock in stocks) \
                    <= bounce_vars[timestamp])

    solver2.Minimize(sum(bounce_vars[timestamp] for timestamp in timestamps[1:]))

    status = solver2.Solve()


    if status == pywraplp.Solver.OPTIMAL:
        print("Task3_C", currency)
        with open(f"task3_C_{currency}.csv", "a", newline= "") as f:
            csvwriter = csv.writer(f)
            csvwriter.writerow([None] + stocks)
            for timestamp in timestamps:
                csvwriter.writerow([timestamp[:7]] + [round(portfolio_vars[timestamp, stock].solution_value() * 100.0, 2) for stock in stocks])
                # print(f"In {timestamp[:7]}")
                # print(", ".join([f"{stock}: {portfolio_vars[timestamp, stock].solution_value() * 100.0:.2f}%" for stock in stocks]))
        
        print("Overall Average Monthly Reward: ", sum(sum(portfolio_vars[timestamp, stock].solution_value() * return_data[timestamp, stock] for stock in stocks) \
                  for timestamp in timestamps[1:]) / (len(timestamps) - 1))
    else:
        print("This problem has no optimal solution")
    
    print("\n\n")
    
if __name__ == '__main__':
    task1()
    task2()
    task3('USD')
    task3('EUR')
    pass
